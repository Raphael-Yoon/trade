# -*- coding: utf-8 -*-
import sys
import os

# Windows 콘솔 UTF-8 설정
if os.name == 'nt':
    os.system('chcp 65001 > nul')
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

sys.path.append(r'C:\Users\newsi\AppData\Roaming\Python\Python313\site-packages')

import OpenDartReader
import pandas as pd
import time
from pykrx import stock
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
import argparse
warnings.filterwarnings('ignore')

# DART API KEY (환경변수로 관리 권장)
API_KEY = "08e04530eea4ba322907021334794e4164002525"

def get_latest_business_day():
    """가장 최근의 영업일을 반환합니다."""
    try:
        end_date = datetime.now().strftime("%Y%m%d")
        start_date = (datetime.now() - timedelta(days=10)).strftime("%Y%m%d")
        ohlcv = stock.get_market_ohlcv(start_date, end_date, "005930")
        if ohlcv.empty:
            return (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
        return ohlcv.index[-1].strftime("%Y%m%d")
    except Exception as e:
        print(f"영업일 조회 오류: {e}. 어제 날짜 사용")
        return (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")

def get_sector_from_pykrx(ticker, sector_df):
    """pykrx 업종 분류 데이터에서 업종 정보를 추출합니다."""
    try:
        if ticker in sector_df.index:
            # 두 번째 컬럼(인덱스 1)이 업종 정보
            sector = sector_df.loc[ticker].iloc[1]
            return sector if pd.notna(sector) else 'N/A'
        return 'N/A'
    except:
        return 'N/A'

def get_dart_financials(dart, ticker, year):
    """OpenDARTReader를 사용하여 재무 데이터를 추출합니다."""
    try:
        # finstate_all 사용 (연결재무제표 우선)
        df = dart.finstate_all(ticker, year)

        if df is None or df.empty:
            return 0, 0, 0, 0

        revenue = 0
        op = 0
        re = 0
        cash = 0

        # 항목명 매칭 (공백 제거 후 'in' 연산자 사용)
        for _, row in df.iterrows():
            acc_name = str(row['account_nm']).replace(" ", "")
            val = pd.to_numeric(row['thstrm_amount'], errors='coerce')
            if pd.isna(val):
                val = 0

            # 매출액
            if acc_name == '매출액' or acc_name == '수익(매출액)':
                if revenue == 0 or acc_name == '매출액':
                    revenue = val
            # 영업이익 (정확히 "영업이익"인 경우만, 영업이익률은 제외)
            elif acc_name == '영업이익':
                op = val
            # 이익잉여금
            elif '이익잉여금' in acc_name and '기타' not in acc_name:
                if re == 0 or acc_name == '이익잉여금':
                    re = val
            # 현금및현금성자산
            elif '현금및현금성자산' in acc_name:
                if cash == 0 or acc_name == '현금및현금성자산':
                    cash = val

        return revenue, op, re, cash
    except Exception as e:
        print(f"[DART] {ticker} 재무제표 조회 실패: {e}")
        return 0, 0, 0, 0

def main(stock_count=100, selected_fields=None):
    try:
        print("=" * 80)
        if stock_count == 0:
            print("KOSPI 전체 종목 데이터 수집 시작")
        else:
            print(f"KOSPI 상위 {stock_count}개 종목 데이터 수집 시작")
        print("수집 우선순위: pykrx > DART API")
        print("=" * 80)

        dart = OpenDartReader(API_KEY)

        # 1. 최근 영업일 조회
        latest_date = get_latest_business_day()
        print(f"\n최근 영업일: {latest_date}")

        # 2. pykrx로 시가총액 상위 100개 종목 조회
        print("\n[1단계] pykrx로 시가총액 및 기본 데이터 수집 중...")
        df_cap = stock.get_market_cap_by_ticker(latest_date, market="KOSPI")
        df_fundamental = stock.get_market_fundamental(latest_date, market="KOSPI")
        df_sector = stock.get_market_sector_classifications(latest_date, market="KOSPI")

        # 상위 N개 종목 (0이면 전체)
        if stock_count == 0:
            df_top100 = df_cap.sort_values(by='시가총액', ascending=False)
            print(f"대상 종목 수: 전체 {len(df_top100)}개")
        else:
            df_top100 = df_cap.sort_values(by='시가총액', ascending=False).head(stock_count)
            print(f"대상 종목 수: {len(df_top100)}개")

        # 3. 데이터 수집 연도 (전년도 사업보고서)
        current_year = datetime.now().year - 1

        results = []
        print(f"\n[2단계] 종목별 상세 데이터 수집 중...")

        # 진행 상황 표시
        total = len(df_top100)
        for idx, ticker in enumerate(df_top100.index, 1):
            name = stock.get_market_ticker_name(ticker)

            # 진행률 표시 (같은 줄에 덮어쓰기)
            print(f"\r진행률: [{idx}/{total}] {idx*100//total}% 완료", end='', flush=True)

            # (1) pykrx 데이터 (최우선)
            pbr = df_fundamental.loc[ticker, 'PBR'] if ticker in df_fundamental.index else 0.0
            per = df_fundamental.loc[ticker, 'PER'] if ticker in df_fundamental.index else 0.0
            eps = df_fundamental.loc[ticker, 'EPS'] if ticker in df_fundamental.index else 0.0
            bps = df_fundamental.loc[ticker, 'BPS'] if ticker in df_fundamental.index else 0.0
            div_yield = df_fundamental.loc[ticker, 'DIV'] if ticker in df_fundamental.index else 0.0

            # ROE 계산 (ROE = EPS / BPS * 100)
            roe = (eps / bps * 100) if bps > 0 else 0.0

            # (2) pykrx에서 업종 정보
            sector = get_sector_from_pykrx(ticker, df_sector)

            # (3) DART API 데이터 (매출액, 영업이익, 이익잉여금, 현금)
            revenue, op, re, cash = get_dart_financials(dart, ticker, current_year)

            # 데이터 저장
            results.append({
                '종목코드': ticker,
                '종목명': name,
                '업종': sector,
                'PBR': round(pbr, 2),
                'PER': round(per, 2),
                'ROE': round(roe, 2),
                'EPS': int(eps),
                'BPS': int(bps),
                '배당수익률': round(div_yield, 2),
                '매출액': int(revenue),
                '영업이익': int(op),
                '이익잉여금': int(re),
                '현금및현금성자산': int(cash)
            })

            time.sleep(0.05)  # API 부하 방지

        print()  # 진행률 표시 후 줄바꿈
        df_result = pd.DataFrame(results)

        # 선택된 필드만 필터링
        if selected_fields:
            # 종목코드와 종목명은 항상 포함
            required_fields = ['종목코드', '종목명']
            fields_to_include = required_fields + [f for f in selected_fields if f not in required_fields]
            # DataFrame에 실제 존재하는 컬럼만 선택
            fields_to_include = [f for f in fields_to_include if f in df_result.columns]
            df_result = df_result[fields_to_include]

        # 4. 엑셀 저장 먼저 수행
        output_file = "result.xlsx"
        df_result.to_excel(output_file, index=False, engine='openpyxl')

        # 엑셀 포맷팅
        wb = load_workbook(output_file)
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        wb.save(output_file)

        # 5. 결과 출력 (인코딩 안전 처리)
        print("\n" + "=" * 80)
        print("[Data Collection Completed]")
        print("=" * 80)
        try:
            print(df_result.head(10).to_string(index=False))
        except:
            pass  # 인코딩 오류 무시

        print(f"\nData saved: {output_file}")
        print(f"Total stocks: {len(df_result)}")

    except Exception as e:
        print(f"\nError occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='KOSPI 상위 종목 데이터 수집')
    parser.add_argument('--count', type=int, default=100,
                       help='수집할 종목 수 (기본값: 100)')
    parser.add_argument('--fields', type=str, default=None,
                       help='수집할 필드 (쉼표로 구분, 예: PBR,PER,ROE)')
    args = parser.parse_args()

    # 필드를 리스트로 변환
    selected_fields = args.fields.split(',') if args.fields else None

    main(stock_count=args.count, selected_fields=selected_fields)
